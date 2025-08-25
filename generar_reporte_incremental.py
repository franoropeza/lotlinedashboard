#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera y actualiza el reporte de billeteras de forma *incremental*:
1) Lee solo los .xls nuevos/modificados de data/
2) Actualiza datasets/movimientos.parquet y manifest.csv
3) Vuelve a crear reporte_movimientos.xlsx y la plantilla con pivots
"""

# ================== IMPORTS Y CONFIG ==================
from pathlib import Path
from datetime import datetime, timedelta
import warnings, shutil, unicodedata, os, time
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.exceptions import InvalidFileException

warnings.filterwarnings("ignore", message="Workbook contains no default style")

## carpetas
ROOT          = Path(__file__).parent
DATA_DIR      = ROOT / "data"        # llegan los .xls diarios
PROC_DIR      = ROOT / "processed"   # .xls ya integrados
DS_DIR        = ROOT / "datasets"    # parquet + manifest
csv_dir       = ROOT / "csv_dashboard" # para CSVs de salida

PROC_DIR.mkdir(exist_ok=True)
DS_DIR.mkdir(exist_ok=True)
csv_dir.mkdir(exist_ok=True)

MANIFEST_FILE = DS_DIR / "manifest.csv"
MASTER_FILE   = DS_DIR / "movimientos.parquet"

## salidas
SALIDA_ANALITICO = ROOT / "reporte_movimientos.xlsx"
TEMPLATE_PATH    = ROOT / "Plantilla-Reporte.xlsm"   # con macros / pivots
SALIDA_XLWINGS   = ROOT / f"ReporteMovimientos-{datetime.now():%Y%m%d}.xlsm"
USUARIOS_FILE    = ROOT / "REPORTE-A-MEDIDA-USUARIOSACTIVOS.xlsx"      # archivo con datos de usuario

SHARE_DIR        = ROOT / "public"                  # carpeta compartida
SHARE_DIR.mkdir(exist_ok=True)
SALIDA_PUBLICA   = SHARE_DIR / "Dashboard-Billeteras.xlsm"
ENABLE_XLWINGS   = True

# hitos
FECHA_LANZ_JUEGOS = pd.Timestamp("2025-04-14")
FECHA_MODO_FULL   = pd.Timestamp("2025-07-07")

# ================== FUNCIONES BASE ==================
def normalizar(txt: str) -> str:
    """Normaliza un texto, eliminando acentos y convirtiendo a minúsculas."""
    if pd.isna(txt): return ""
    txt = unicodedata.normalize("NFKD", str(txt))
    return "".join(c for c in txt if not unicodedata.combining(c)).lower()

def leer_movimientos(archivo: Path) -> pd.DataFrame | None:
    """Lee un .xls y devuelve un dataframe normalizado (o None si no detecta encabezado)."""
    crudo = pd.read_excel(archivo, header=None)
    header_mask = crudo.apply(
        lambda fila: fila.apply(normalizar).str.contains("tipo mov", na=False).any(),
        axis=1,
    )
    if not header_mask.any():
        print(f"⚠️  Encabezado no encontrado en {archivo.name} — omitido")
        return None
    header_idx = header_mask.idxmax()

    df = pd.read_excel(
        archivo, header=header_idx,
        usecols=["Nro. Transacción","Fecha","Tipo Mov.","Documento","Movimiento","Importe"]
    )
    df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors="coerce")
    df["Importe"] = (
        df["Importe"].astype(str)
                     .str.replace(".", "", regex=False)
                     .str.replace(",", ".", regex=False)
                     .astype(float)
    )
    return df

def clasificar_canal(movimiento: str, tipo_mov: str = "") -> str:
    """
    Clasifica un movimiento como MODO, Retail (TJ/Agencia) u Otro.
    Se apoya tanto en 'Movimiento' como en 'Tipo Mov.'.
    """
    mov_norm = normalizar(movimiento)
    tipo_norm = normalizar(tipo_mov)

    # Detectar MODO
    if "modo" in mov_norm or "modo" in tipo_norm:
        return "MODO"

    # Detectar Tarjeta/TJ (se mapea como Retail porque así espera el dashboard)
    if "tj" in mov_norm or "tarjeta" in mov_norm or "tj" in tipo_norm or "tarjeta" in tipo_norm:
        return "Retail"

    # Detectar Agencia/POS/Caja → también Retail
    if any(k in mov_norm for k in ["agencia", "pos", "caja"]):
        return "Retail"

    # Fallback
    return "Otro"


def get_mtime(p: Path) -> int:
    """Obtiene el tiempo de última modificación de un archivo."""
    return int(p.stat().st_mtime)

# ================== BLOQUE INCREMENTAL ==================
print("🔄  Paso 1: identificar archivos nuevos…")
# robust manifest loading
expected_cols = ["archivo", "mod_time"]
if MANIFEST_FILE.exists():
    manifest = pd.read_csv(MANIFEST_FILE)
    if list(manifest.columns) != expected_cols:
        print(f"⚠️  Manifest file '{MANIFEST_FILE}' has incorrect columns. A new one will be created.")
        manifest = pd.DataFrame(columns=expected_cols)
else:
    manifest = pd.DataFrame(columns=expected_cols)

pendientes = []
for f in DATA_DIR.glob("*.xls"):
    mt = get_mtime(f)
    fila = manifest.loc[manifest["archivo"] == f.name]
    if fila.empty or fila.iloc[0]["mod_time"] != mt:
        pendientes.append((f, mt))

nuevos_df = []
if pendientes:
    for f, mt in pendientes:
        print("   • Procesando", f.name)
        df_tmp = leer_movimientos(f)
        if df_tmp is not None:
            nuevos_df.append(df_tmp)
            # mover a processed/YYYY-MM/
            dest_dir = PROC_DIR / f"{df_tmp['Fecha'].dt.to_period('M').iloc[0]}"
            dest_dir.mkdir(exist_ok=True)
            try:
                shutil.move(str(f), dest_dir / f.name)
            except shutil.Error as e:
                print(f"⚠️ Error al mover el archivo {f.name}: {e}. El archivo puede ya existir en el destino.")
            # actualizar manifest
            manifest = manifest[manifest["archivo"] != f.name]
            manifest.loc[len(manifest)] = [f.name, mt]

    # actualizar parquet maestro
    df_new = pd.concat(nuevos_df, ignore_index=True)
    if MASTER_FILE.exists():
        df_old = pd.read_parquet(MASTER_FILE)
        data_total = pd.concat([df_old, df_new], ignore_index=True)
        # CRITICAL FIX: keep='last' to ensure updates are reflected
        data_total.drop_duplicates(subset=["Nro. Transacción"], keep='last', inplace=True)
    else:
        data_total = df_new
    data_total.to_parquet(MASTER_FILE, index=False)
    manifest.to_csv(MANIFEST_FILE, index=False)
    print(f"✅ Agregados {len(df_new)} movimientos nuevos.")
else:
    if not MASTER_FILE.exists():
        raise RuntimeError("No hay parquet maestro y no se encontraron .xls para procesar.")
    print("✅ Dataset maestro ya estaba al día.")

# ================== BLOQUE DE GENERACIÓN DE REPORTE ==================
# **Asegurarse de cargar la data más reciente**
try:
    data_total = pd.read_parquet(MASTER_FILE)
    print(f"📊  Movimientos totales en dataset: {len(data_total):,}")
    data = data_total.copy()
except Exception as e:
    print(f"❌ Error al cargar el dataset maestro: {e}")
    exit()

## ========= APUESTAS =========
apuestas = data[data["Tipo Mov."].str.contains("apuesta|jugada", case=False, na=False)].copy()
apuestas["AñoMes"]    = apuestas["Fecha"].dt.to_period("M")
apuestas["Fecha_Dia"] = apuestas["Fecha"].dt.date
apuestas["Juego"] = (
    apuestas["Movimiento"].str.replace(r"(?i)jugada\s*-\s*", "", regex=True).str.strip()
)
apuestas["Juego_norm"] = apuestas["Juego"].apply(normalizar)

dias_map = {0:"Lunes",1:"Martes",2:"Miércoles",3:"Jueves",4:"Viernes",5:"Sábado",6:"Domingo"}
apuestas["Dia_Sem"] = apuestas["Fecha"].dt.weekday.map(dias_map)

# ========= CLIENTE_MES =========
cliente_mes = (
    apuestas.groupby(["Documento", "AñoMes", "Juego"])
            .agg(Bets=("Importe", "count"), Gastado=("Importe", "sum"))
            .reset_index()
)

# ========= TOP GAMES =========
top_games_total = (
    apuestas.groupby("Juego")
            .agg(Bets_Totales=("Importe", "count"),
                 Jugadores_Unicos=("Documento", "nunique"),
                 Gastado=("Importe", "sum"))
            .sort_values("Bets_Totales", ascending=False)
            .reset_index()
)

top_games_mes = (
    apuestas.groupby(["AñoMes", "Juego"])
            .agg(Bets_Mes=("Importe", "count"),
                 Jugadores_Mes=("Documento", "nunique"),
                 Gastado_Mes=("Importe", "sum"))
            .reset_index()
            .sort_values(["AñoMes", "Bets_Mes"], ascending=[False, False])
)

# ========= PESTAÑAS POR JUEGO =========
GAME_PATTERNS = {
    "Tombo_Express": r"tombo express",
    "Tombola":       r"t(?:o|ó)mbola",
    "Quini6":        r"quini\s*6",
    "Loto_Plus":     r"loto(?:\s*plus)?",
}
game_summaries = {}
for sheet, pattern in GAME_PATTERNS.items():
    mask = apuestas["Juego_norm"].str.contains(pattern, regex=True, na=False)
    tmp = apuestas.loc[mask]
    summary = (
        tmp.groupby("Documento")
           .agg(Bets=("Importe", "count"), Gastado=("Importe", "sum"))
           .sort_values("Gastado", ascending=False)
           .reset_index()
    )
    game_summaries[sheet] = summary

# ========= RECARGAS / RETIROS / PREMIOS =========
# Tomamos SOLO verdaderas cargas de saldo desde MODO o TJ (no "depósitos" genéricos ni bonificaciones)

# Normalizar textos (asume que ya definiste normalizar(); si no, avisá y te lo agrego)
tm_norm = data["Tipo Mov."].apply(normalizar)

# Filtra "Carga saldo desde MODO/TJ" (soporta variantes de espacios)
mask_carga = tm_norm.str.match(r"^carga\s+saldo\s+desde\s*(?:modo|tj)\b", na=False)

cargas = data.loc[mask_carga].copy()

# Canal leyendo Movimiento y Tipo Mov. (TJ -> Retail, MODO -> MODO)
cargas["Canal"] = cargas.apply(
    lambda r: clasificar_canal(r.get("Movimiento", ""), r.get("Tipo Mov.", "")),
    axis=1
)

# Metodo: lo que espera tu dashboard (todo lo que no sea MODO va como Retail)
cargas["Metodo"] = np.where(cargas["Canal"] == "MODO", "MODO", "Retail")

# Derivados de fecha/hora
cargas["Fecha_Dia"] = pd.to_datetime(cargas["Fecha"], errors="coerce").dt.date
cargas["Hora"] = pd.to_datetime(cargas["Fecha"], errors="coerce").dt.hour

# Agregación diaria por canal
recargas_diario_canal = (
    cargas.groupby(["Fecha_Dia", "Canal"], as_index=False)
          .agg(Recargas=("Importe", "count"),
               Monto=("Importe", "sum"),
               Usuarios_Unicos=("Documento", "nunique"))
)

recargas_dia_monto = (
    recargas_diario_canal.pivot(index="Fecha_Dia", columns="Canal", values="Monto")
                         .fillna(0.0)
                         .reset_index()
                         .sort_values("Fecha_Dia")
)
recargas_dia_cant = (
    recargas_diario_canal.pivot(index="Fecha_Dia", columns="Canal", values="Recargas")
                         .fillna(0)
                         .reset_index()
                         .sort_values("Fecha_Dia")
)

modo_diario = (
    cargas.loc[cargas["Canal"] == "MODO"]
          .groupby("Fecha_Dia")
          .agg(
              Recargas_MODO   = ("Importe", "count"),
              Monto_MODO      = ("Importe", "sum"),
              Usuarios_Unicos = ("Documento", "nunique"),
          )
          .reset_index()
          .sort_values("Fecha_Dia", ascending=False)
)

mask_retiro = data["Tipo Mov."].str.contains("retiro|transferencia salida", case=False, na=False)
retiros = data.loc[mask_retiro].copy()
retiros["Fecha_Dia"] = retiros["Fecha"].dt.date
retiros_diario = (
    retiros.groupby("Fecha_Dia")
           .agg(Retiros=("Importe", "count"),
                Monto_Retirado=("Importe", "sum"),
                Clientes_Unicos=("Documento", "nunique"))
           .reset_index()
           .sort_values("Fecha_Dia", ascending=False)
)

mask_premio = data["Tipo Mov."].str.contains("premio", case=False, na=False)
premios = data.loc[mask_premio].copy()
premios_resumen = (
    premios.groupby("Documento")
           .agg(Premios_Cobrados=("Importe", "count"), Monto_Premios=("Importe", "sum"))
           .sort_values("Monto_Premios", ascending=False)
           .reset_index()
)

# ========= Juego por día (detalle) =========
juego_dia_detalle = (
    apuestas.groupby(["Fecha_Dia", "Juego"])
            .agg(Bets=("Importe", "count"),
                 Usuarios_Unicos_Dia=("Documento", "nunique"),
                 Gastado_Dia=("Importe", "sum"))
            .reset_index()
            .sort_values(["Fecha_Dia", "Juego"])
)

dia_totales = (
    apuestas.groupby("Dia_Sem")
            .agg(Bets=("Importe", "count"),
                 Usuarios_Unicos=("Documento", "nunique"))
            .reset_index()
)

# ========= Retención MODO =========
primer_mov_total = (
    data.sort_values("Fecha")
        .groupby("Documento", as_index=False)
        .first()
        .rename(columns={"Fecha": "Fecha_PrimerMov"})
)

modo_all = cargas.loc[cargas["Canal"] == "MODO"].copy()
primera_modo = (
    modo_all.sort_values("Fecha")
            .groupby("Documento", as_index=False)
            .first()
            .rename(columns={"Fecha": "Fecha_Corte"})
)
retencion_base = primera_modo.merge(primer_mov_total, on="Documento", how="left")
retencion_base["Es_Nuevo"] = retencion_base["Fecha_PrimerMov"] == retencion_base["Fecha_Corte"]

apuestas_only = apuestas[["Documento", "Fecha"]].copy()
joined = retencion_base[["Documento", "Fecha_Corte"]].merge(apuestas_only, on="Documento", how="left")
joined["Posterior"]     = joined["Fecha"] >  joined["Fecha_Corte"]
joined["Dia_Siguiente"] = (joined["Fecha"] > joined["Fecha_Corte"]) & \
                          (joined["Fecha"] <= joined["Fecha_Corte"] + pd.Timedelta(days=1))
joined["Mes_Siguiente"] = (joined["Fecha"] > joined["Fecha_Corte"]) & \
                          (joined["Fecha"] <= joined["Fecha_Corte"] + pd.Timedelta(days=30))

flags = (
    joined.groupby("Documento")
          .agg(Jugo_Posterior=("Posterior", "any"),
               Jugo_Dia_Sig=("Dia_Siguiente", "any"),
               Jugo_Mes_Sig=("Mes_Siguiente", "any"))
          .reset_index()
)
retencion_modo = retencion_base.merge(flags, on="Documento", how="left").fillna(False)
print(f"▶︎ Cantidad de usuarios NUEVOS que cargaron con MODO: {int(retencion_modo['Es_Nuevo'].sum())}")

# ========= Crecimiento / Hitos =========
primer_mov_total["AñoMes_PrimerMov"] = primer_mov_total["Fecha_PrimerMov"].dt.to_period("M")
usuarios_mes = (
    primer_mov_total.groupby("AñoMes_PrimerMov")
                    .agg(Nuevos=("Documento", "nunique"))
                    .reset_index()
                    .sort_values("AñoMes_PrimerMov")
)
usuarios_mes["Acumulado"] = usuarios_mes["Nuevos"].cumsum()
activos_mes = (
    apuestas.groupby(apuestas["Fecha"].dt.to_period("M"))["Documento"]
            .nunique()
            .reset_index(name="Jugadores_Activos_Mes")
            .rename(columns={"Fecha": "AñoMes"})
)
activos_mes["AñoMes"] = activos_mes["AñoMes"].astype(str)
usuarios_mes["AñoMes"] = usuarios_mes["AñoMes_PrimerMov"].astype(str)
usuarios_mes = usuarios_mes.merge(activos_mes, on="AñoMes", how="left")
usuarios_mes = usuarios_mes[["AñoMes", "Nuevos", "Acumulado", "Jugadores_Activos_Mes"]]

nuevos_desde_juegos = primer_mov_total.loc[
    primer_mov_total["Fecha_PrimerMov"] >= FECHA_LANZ_JUEGOS, "Documento"
].nunique()
jugadores_desde_juegos = apuestas.loc[
    apuestas["Fecha"] >= FECHA_LANZ_JUEGOS, "Documento"
].nunique()
mask_newgames = (
    (apuestas["Fecha"] >= FECHA_LANZ_JUEGOS) &
    apuestas["Juego_norm"].str.contains(r"(?:quini\s*6|loto(?:\s*plus)?)", regex=True, na=False)
)
jugadores_quini_loto_desde = apuestas.loc[mask_newgames, "Documento"].nunique()

modo_post_docs = set(
    cargas.loc[(cargas["Fecha"] >= FECHA_MODO_FULL) & (cargas["Canal"] == "MODO"), "Documento"].unique()
)
usuarios_modo_desde = len(modo_post_docs)
jugadores_post_modo_docs = set(apuestas.loc[apuestas["Fecha"] >= FECHA_MODO_FULL, "Documento"].unique())
jugadores_post_modo = len(jugadores_post_modo_docs)
jugadores_modo_y_jugaron = len(modo_post_docs & jugadores_post_modo_docs)

usuarios_hitos = pd.DataFrame({
    "Concepto": [
        f"Nuevos >= {FECHA_LANZ_JUEGOS.strftime('%Y-%m-%d')} (cualquier mov.)",
        f"Jugadores >= {FECHA_LANZ_JUEGOS.strftime('%Y-%m-%d')} (cualquier apuesta)",
        f"Apostaron Quini/Loto >= {FECHA_LANZ_JUEGOS.strftime('%Y-%m-%d')}",
        f"Recargaron MODO >= {FECHA_MODO_FULL.strftime('%Y-%m-%d')}",
        f"Jugadores >= {FECHA_MODO_FULL.strftime('%Y-%m-%d')} (cualquier apuesta)",
        f"Recargaron MODO y jugaron >= {FECHA_MODO_FULL.strftime('%Y-%m-%d')}",
    ],
    "Valor": [
        nuevos_desde_juegos,
        jugadores_desde_juegos,
        jugadores_quini_loto_desde,
        usuarios_modo_desde,
        jugadores_post_modo,
        jugadores_modo_y_jugaron,
    ],
})

# ========= Comparativa Before/After 07/07 =========
dep_before = cargas.loc[cargas["Fecha"] < FECHA_MODO_FULL, "Importe"].sum()
dep_after  = cargas.loc[cargas["Fecha"] >= FECHA_MODO_FULL, "Importe"].sum()
rec_before = apuestas.loc[apuestas["Fecha"] < FECHA_MODO_FULL, "Importe"].sum()
rec_after  = apuestas.loc[apuestas["Fecha"] >= FECHA_MODO_FULL, "Importe"].sum()

comparativa_modo = pd.DataFrame({
    "Periodo": [f"Before {FECHA_MODO_FULL.strftime('%d/%m/%Y')}", f"After {FECHA_MODO_FULL.strftime('%d/%m/%Y')}"],
    "Depositos_$": [dep_before, dep_after],
    "Recaudacion_$": [rec_before, rec_after],
})

# ========= KPIs Resumen =========
promedio_deposito    = cargas["Importe"].mean() if len(cargas) else 0.0
cant_unicos_total    = data["Documento"].nunique()
cant_unicos_apuestan = apuestas["Documento"].nunique()
cant_recargas_unicas = cargas["Documento"].nunique()

agg_canal = (
    cargas.groupby("Canal")
          .agg(Recargas=("Importe", "count"),
               Usuarios_Unicos=("Documento", "nunique"),
               Monto=("Importe", "sum"))
          .reset_index()
)
recargas_modo   = int(agg_canal.loc[agg_canal["Canal"]=="MODO","Recargas"].sum()) if not agg_canal.empty else 0
recargas_retail = int(agg_canal.loc[agg_canal["Canal"]=="Retail","Recargas"].sum()) if not agg_canal.empty else 0
monto_modo      = float(agg_canal.loc[agg_canal["Canal"]=="MODO","Monto"].sum()) if not agg_canal.empty else 0.0
monto_retail    = float(agg_canal.loc[agg_canal["Canal"]=="Retail","Monto"].sum()) if not agg_canal.empty else 0.0

resumen_kpis = pd.DataFrame({
    "KPI": [
        "Promedio depósito $",
        "Usuarios únicos (cualquier mov.)",
        "Usuarios únicos apostadores",
        "Usuarios únicos que recargaron",
        "Recargas - MODO",
        "Recargas - Retail",
        "Monto MODO $",
        "Monto Retail $",
    ],
    "Valor": [
        promedio_deposito,
        cant_unicos_total,
        cant_unicos_apuestan,
        cant_recargas_unicas,
        recargas_modo,
        recargas_retail,
        monto_modo,
        monto_retail,
    ]
})

# Inicializar DataFrames para evitar NameError
top10_contactos = None
usuarios_nuevos_modo = None
usuarios_reactivados_modo = None
top10_por_juego_con_datos = {}


# ========= Top10 contactos (opcional) =========
if USUARIOS_FILE.exists():
    try:
        usuarios_raw = pd.read_excel(USUARIOS_FILE)
        usuarios_raw.columns = usuarios_raw.columns.str.strip()
        if "Fecha_Alta" not in usuarios_raw.columns and "Fecha Alta" in usuarios_raw.columns:
            usuarios_raw.rename(columns={"Fecha Alta": "Fecha_Alta"}, inplace=True)
        usuarios_raw["Fecha_Alta"] = pd.to_datetime(usuarios_raw["Fecha_Alta"], dayfirst=True, errors="coerce")
        usuarios = usuarios_raw.copy()

        if "DNI" not in usuarios.columns:
            cand = [c for c in usuarios.columns if "dni" in c.lower() or "doc" in c.lower()]
            if cand:
                usuarios.rename(columns={cand[0]: "DNI"}, inplace=True)
            else:
                raise KeyError("No se encontró columna DNI en usuarios_limpio.csv")
        usuarios["DNI"] = pd.to_numeric(usuarios["DNI"], errors="coerce")
        usuarios = usuarios.dropna(subset=["DNI"]).drop_duplicates(subset="DNI", keep="last")

        jugadas_por_doc = (
            apuestas.groupby("Documento")
                    .agg(Bets_Total=("Importe", "count"),
                         Gastado_Total=("Importe", "sum"))
                    .reset_index()
        )

        top10_contactos = (
            jugadas_por_doc.merge(usuarios, how="inner", left_on="Documento", right_on="DNI")
                           .sort_values("Bets_Total", ascending=False)
                           .head(10)
        )
    except Exception as e:
        print(f"⚠️  No se pudo generar Top10_Contactos: {e}")

# ==============================================
# Detectar usuarios NUEVOS y REACTIVADOS con MODO (desde 7/7)
# ==============================================
try:
    modo_movimientos = cargas[cargas["Canal"] == "MODO"][["Documento", "Fecha", "Importe"]].copy()
    modo_movimientos.to_csv(csv_dir / "movimientos_modo.csv", index=False)
    print("✅ CSV generado: movimientos_modo.csv")
except Exception as e:
    print(f"⚠️ Error al generar movimientos_modo.csv: {e}")

if USUARIOS_FILE.exists():
    try:
        usuarios = pd.read_excel(USUARIOS_FILE)
        usuarios.columns = usuarios.columns.str.strip()
        if "Fecha_Alta" not in usuarios.columns and "Fecha Alta" in usuarios.columns:
            usuarios.rename(columns={"Fecha Alta": "Fecha_Alta"}, inplace=True)
        usuarios["Fecha_Alta"] = pd.to_datetime(usuarios["Fecha_Alta"], dayfirst=True, errors="coerce")
        
        usuarios["Documento"] = pd.to_numeric(
            usuarios["Documento"].astype(str).str.extract(r"(\d+)")[0].str.lstrip("0"),
            errors="coerce"
        )
        usuarios = usuarios.dropna(subset=["Documento"])

        usuarios_nuevos_modo = usuarios[usuarios["Fecha_Alta"] >= FECHA_MODO_FULL].copy()
        usuarios_nuevos_modo = usuarios_nuevos_modo[["Documento", "Fecha_Alta", "Usuario", "Correo"]].copy()
        
        print("✅ DataFrame generado: usuarios_nuevos_modo")

        antiguos = usuarios[usuarios["Fecha_Alta"].dt.year.between(2021, 2024)].copy()

        cargas_modo = pd.read_csv(csv_dir / "movimientos_modo.csv")
        cargas_modo["Fecha"] = pd.to_datetime(cargas_modo["Fecha"], dayfirst=True, errors="coerce")
        cargas_modo["Documento"] = pd.to_numeric(
            cargas_modo["Documento"].astype(str).str.extract(r"(\d+)")[0].str.lstrip("0"),
            errors="coerce"
        )

        jugaban_antes = set(cargas_modo[cargas_modo["Fecha"] < FECHA_MODO_FULL]["Documento"])
        jugaban_despues = set(cargas_modo[cargas_modo["Fecha"] >= FECHA_MODO_FULL]["Documento"])

        reactivados = antiguos[antiguos["Documento"].isin(jugaban_despues - jugaban_antes)].copy()
        
        primer_modo = cargas_modo[cargas_modo["Fecha"] >= FECHA_MODO_FULL].sort_values("Fecha").drop_duplicates("Documento")
        usuarios_reactivados_modo = reactivados.merge(
            primer_modo[["Documento", "Fecha"]],
            left_on="Documento", right_on="Documento", how="left"
        )
        
        reactivados_csv_path = csv_dir / "reactivados_modo.csv"
        if reactivados_csv_path.exists():
            os.remove(reactivados_csv_path)

        usuarios_reactivados_modo.to_csv(reactivados_csv_path, index=False)
        print("✅ CSV generado: reactivados_modo.csv")

        nuevos_csv_path = csv_dir / "nuevos_modo.csv"
        if nuevos_csv_path.exists():
            os.remove(nuevos_csv_path)
        usuarios_nuevos_modo.to_csv(nuevos_csv_path, index=False)
        print("✅ CSV generado: nuevos_modo.csv")

        total_nuevos_modo = usuarios_nuevos_modo.shape[0]
        pd.DataFrame([{"KPI": f"Total Nuevos Usuarios desde {FECHA_MODO_FULL.strftime('%d/%m/%Y')}", "Valor": total_nuevos_modo}]).to_csv(csv_dir / "total_usuarios_nuevos_modo.csv", index=False)
        print("✅ CSV generado: total_usuarios_nuevos_modo.csv")


    except Exception as e:
        print(f"⚠️ Error al detectar usuarios reactivados o nuevos por MODO: {e}")

# ========= Top10 por juego con datos personales =========
if USUARIOS_FILE.exists():
    try:
        usuarios = pd.read_excel(USUARIOS_FILE)
        usuarios.columns = usuarios.columns.str.strip()
        if "Fecha_Alta" not in usuarios.columns and "Fecha Alta" in usuarios.columns:
            usuarios.rename(columns={"Fecha Alta": "Fecha_Alta"}, inplace=True)
        usuarios["Fecha_Alta"] = pd.to_datetime(usuarios["Fecha_Alta"], dayfirst=True, errors="coerce")

        if "Documento" in usuarios.columns:
            usuarios.rename(columns={"Documento": "DNI"}, inplace=True)

        usuarios["DNI"] = pd.to_numeric(usuarios["DNI"], errors="coerce").astype("Int64")
        usuarios = usuarios.dropna(subset=["DNI"])[["DNI", "Usuario", "Correo"]].drop_duplicates("DNI")

        for juego, df in game_summaries.items():
            df = df.copy()
            df["Documento"] = pd.to_numeric(df["Documento"], errors="coerce").astype("Int64")

            merged = df.merge(usuarios, left_on="Documento", right_on="DNI", how="inner")
            top10 = merged.sort_values("Gastado", ascending=False).head(10)

            top10_por_juego_con_datos[juego] = top10

    except Exception as e:
        print(f"⚠️  Error al generar Top10 por juego con datos: {e}")

# ========= KPIs de Actividad de Usuarios (para Upgrade 1 Dashboard) =========
if USUARIOS_FILE.exists():
    try:
        total_registrados = len(usuarios)
        total_activos = data["Documento"].nunique()
        total_inactivos = total_registrados - total_activos
        tasa_actividad = (total_activos / total_registrados) * 100 if total_registrados > 0 else 0

        nuevos_kpis = pd.DataFrame({
            "KPI": [
                "Total Usuarios Registrados",
                "Usuarios Activos (con mov.)",
                "Usuarios Inactivos (sin mov.)",
                "Tasa de Actividad (%)"
            ],
            "Valor": [
                total_registrados,
                total_activos,
                total_inactivos,
                tasa_actividad
            ]
        })
        resumen_kpis = pd.concat([resumen_kpis, nuevos_kpis], ignore_index=True)
        print("✅ KPIs de actividad de usuarios calculados.")
    except NameError:
        print("⚠️  DataFrame 'usuarios' no fue creado, no se pueden calcular KPIs de actividad.")
    except Exception as e:
        print(f"⚠️ Error al calcular KPIs de actividad de usuarios: {e}")

# ========= Carga de datos de usuarios (con FIX para Fecha_Alta) =========
usuarios = pd.DataFrame() # Inicializar como DataFrame vacío
if USUARIOS_FILE.exists():
    try:
        usuarios_raw = pd.read_excel(USUARIOS_FILE)
        usuarios_raw.columns = [str(c).strip() for c in usuarios_raw.columns]
        
        # --- FIX INTELIGENTE PARA 'Fecha_Alta' ---
        fecha_alta_col = None
        for col in usuarios_raw.columns:
            norm_col = normalizar(col)
            if 'fecha' in norm_col and 'alta' in norm_col:
                fecha_alta_col = col
                break
        
        if fecha_alta_col:
            usuarios_raw.rename(columns={fecha_alta_col: "Fecha_Alta"}, inplace=True)
            usuarios_raw["Fecha_Alta"] = pd.to_datetime(usuarios_raw["Fecha_Alta"], dayfirst=True, errors="coerce")
        else:
            print("⚠️ ADVERTENCIA: No se encontró una columna 'Fecha_Alta' en el archivo de usuarios. El análisis de retención no funcionará.")

        # --- FIX INTELIGENTE PARA 'DNI' / 'Documento' ---
        dni_col = None
        if "DNI" in usuarios_raw.columns:
            dni_col = "DNI"
        elif "Documento" in usuarios_raw.columns:
            dni_col = "Documento"
        else:
            for col in usuarios_raw.columns:
                norm_col = normalizar(col)
                if 'dni' in norm_col or 'documento' in norm_col:
                    dni_col = col
                    break
        
        if dni_col:
            usuarios_raw.rename(columns={dni_col: "DNI"}, inplace=True)
            usuarios_raw["DNI"] = pd.to_numeric(usuarios_raw["DNI"], errors="coerce")
            usuarios = usuarios_raw.dropna(subset=["DNI"]).drop_duplicates(subset="DNI", keep="last").copy()
            print("✅ Archivo de usuarios cargado y procesado.")
        else:
            print("⚠️ ADVERTENCIA: No se encontró una columna 'DNI' o 'Documento' en el archivo de usuarios.")

    except Exception as e:
        print(f"⚠️ No se pudo procesar el archivo de usuarios: {e}")
        
# ========= Análisis de Retención de Nuevos Usuarios (UPGRADED) =========
print("🔄 Calculando análisis de retención de usuarios mejorado...")
if not usuarios.empty and "Fecha_Alta" in usuarios.columns:
    try:
        # 1. Obtener la fecha de alta de cada usuario
        df_base_usuarios = usuarios[["DNI", "Fecha_Alta"]].copy()
        df_base_usuarios.rename(columns={"DNI": "Documento"}, inplace=True)
        df_base_usuarios["Cohorte_Mes"] = df_base_usuarios["Fecha_Alta"].dt.to_period("M")

        # 2. Encontrar la fecha de la primera apuesta de cada usuario
        primera_apuesta = apuestas.groupby("Documento")["Fecha"].min().reset_index()
        primera_apuesta.rename(columns={"Fecha": "Fecha_Primera_Apuesta"}, inplace=True)
        
        # 3. Unir la fecha de alta con la primera apuesta
        df_funnel = pd.merge(df_base_usuarios, primera_apuesta, on="Documento", how="left")
        df_funnel.dropna(subset=["Fecha_Primera_Apuesta"], inplace=True) # Solo usuarios que han apostado al menos una vez

        # 4. Encontrar la primera recarga DESPUÉS de la primera apuesta
        cargas_posteriores = pd.merge(cargas, df_funnel[['Documento', 'Fecha_Primera_Apuesta']], on="Documento")
        cargas_posteriores = cargas_posteriores[cargas_posteriores["Fecha"] > cargas_posteriores["Fecha_Primera_Apuesta"]]
        primera_recarga_posterior = cargas_posteriores.groupby("Documento")["Fecha"].min().reset_index()
        primera_recarga_posterior.rename(columns={"Fecha": "Fecha_Recarga_Posterior"}, inplace=True)

        # 5. Encontrar la primera apuesta DESPUÉS de la primera apuesta (es decir, la segunda apuesta)
        apuestas_posteriores = pd.merge(apuestas, df_funnel[['Documento', 'Fecha_Primera_Apuesta']], on="Documento")
        apuestas_posteriores = apuestas_posteriores[apuestas_posteriores["Fecha"] > apuestas_posteriores["Fecha_Primera_Apuesta"]]
        segunda_apuesta = apuestas_posteriores.groupby("Documento")["Fecha"].min().reset_index()
        segunda_apuesta.rename(columns={"Fecha": "Fecha_Segunda_Apuesta"}, inplace=True)

        # 6. Marcar usuarios retenidos: aquellos que recargaron Y volvieron a apostar
        df_funnel = pd.merge(df_funnel, primera_recarga_posterior, on="Documento", how="left")
        df_funnel = pd.merge(df_funnel, segunda_apuesta, on="Documento", how="left")
        df_funnel["Retenido"] = ~df_funnel["Fecha_Recarga_Posterior"].isna() & ~df_funnel["Fecha_Segunda_Apuesta"].isna()

        # 7. Calcular el resumen por cohorte
        cohort_summary = df_funnel.groupby("Cohorte_Mes").agg(
            Total_Nuevos_Usuarios=("Documento", "nunique"),
            Retenidos=("Retenido", "sum")
        ).reset_index()
        
        # Renombrar columnas para compatibilidad con el dashboard existente
        cohort_summary.rename(columns={
            "Retenidos": "Retenidos_30_Dias",
        }, inplace=True)

        cohort_summary["Tasa_Retencion_30_Dias"] = (cohort_summary["Retenidos_30_Dias"] / cohort_summary["Total_Nuevos_Usuarios"]) * 100
        # Añadimos una columna dummy para la de 7 días para no romper el gráfico
        cohort_summary["Tasa_Retencion_7_Dias"] = np.nan 

        cohort_summary["Cohorte_Mes"] = cohort_summary["Cohorte_Mes"].astype(str)
        cohort_summary.to_csv(csv_dir / "retencion_cohorts.csv", index=False)
        print("✅ CSV generado: retencion_cohorts.csv con lógica de retención mejorada.")

    except Exception as e:
        print(f"⚠️ Error al generar el análisis de retención mejorado: {e}")
else:
    print("ℹ️  Saltando análisis de retención porque los datos de usuarios no están disponibles o no tienen 'Fecha_Alta'.")




# ================== Exportar datos para Dashboard Dinámico ==================
# Exportar un log detallado de apuestas con datos de usuario para el Top 10 dinámico
if USUARIOS_FILE.exists():
    try:
        # Re-usar el dataframe 'usuarios' ya cargado y limpiado
        apuestas_con_usuarios = apuestas.merge(
            usuarios,
            how="inner",
            left_on="Documento",
            right_on="DNI"
        )
        # Seleccionar columnas relevantes para no exponer datos innecesarios
        columnas_exportar = [
            "Fecha", "Documento", "Usuario", "Correo", "Juego", "Importe"
        ]
        apuestas_con_usuarios[columnas_exportar].to_csv(
            csv_dir / "apuestas_con_usuarios.csv", index=False
        )
        print("✅ CSV generado: apuestas_con_usuarios.csv para Top 10 dinámico.")
    except NameError:
        print("⚠️  DataFrame 'usuarios' no fue creado, no se puede generar apuestas_con_usuarios.csv.")
    except Exception as e:
        print(f"⚠️ Error al generar apuestas_con_usuarios.csv: {e}")

# Exportar apuestas agregadas por día para cálculo de recaudación
apuestas_diario = (
    apuestas.groupby(apuestas["Fecha"].dt.date)
            .agg(Recaudacion=("Importe", "sum"))
            .reset_index()
            .rename(columns={"Fecha": "Fecha_Dia"})
)
apuestas_diario.to_csv(csv_dir / "apuestas_diario.csv", index=False)
print("✅ CSV generado: apuestas_diario.csv para KPI de recaudación.")


# Nuevo: generar CSV para el total de apuestas por juego y por mes
try:
    total_juegos_mes = (
        apuestas.groupby(["AñoMes", "Juego"])["Importe"]
        .count()
        .rename("Total_Bets")
        .reset_index()
    )
    total_juegos_mes.to_csv(csv_dir / "total_juegos_mes.csv", index=False)
    print("✅ CSV generado: total_juegos_mes.csv")
except Exception as e:
    print(f"⚠️ Error al generar total_juegos_mes.csv: {e}")


# ========= Exportar: archivo analítico con gráficos (openpyxl) =========
print(f"📝 Generando el archivo analítico: {SALIDA_ANALITICO}...")
try:
    with pd.ExcelWriter(SALIDA_ANALITICO, engine="openpyxl", mode="w") as writer:
        resumen_kpis.to_excel(writer,              sheet_name="Resumen_Datos",     index=False)
        cliente_mes.to_excel(writer,               sheet_name="Cliente_Mes",       index=False)
        top_games_total.to_excel(writer,           sheet_name="Top_Games_Total",   index=False)
        top_games_mes.to_excel(writer,             sheet_name="Top_Games_Mes",     index=False)

        for sheet, df in game_summaries.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

        recargas_diario_canal.to_excel(writer,     sheet_name="Recargas_Diario",   index=False)
        recargas_dia_monto.to_excel(writer,        sheet_name="Recargas_Dia_Monto",index=False)
        recargas_dia_cant.to_excel(writer,         sheet_name="Recargas_Dia_Cant", index=False)
        modo_diario.to_excel(writer,               sheet_name="MODO_Diario",       index=False)
        retiros_diario.to_excel(writer,            sheet_name="Retiros_Diario",    index=False)
        premios_resumen.to_excel(writer,           sheet_name="Ganadores",         index=False)
        juego_dia_detalle.to_excel(writer,         sheet_name="Juego_Dia_Detalle", index=False)
        dia_totales.to_excel(writer,               sheet_name="Dia_Totales",       index=False)
        retencion_modo.to_excel(writer,            sheet_name="Retencion_MODO",    index=False)
        usuarios_mes.to_excel(writer,              sheet_name="Usuarios_Mes",      index=False)
        usuarios_hitos.to_excel(writer,            sheet_name="Usuarios_Hitos",    index=False)
        comparativa_modo.to_excel(writer,          sheet_name="Comparativa_MODO",  index=False)

        if top10_contactos is not None and not top10_contactos.empty:
            top10_contactos.to_excel(writer,      sheet_name="Top10_Contactos",    index=False)
        
        if usuarios_nuevos_modo is not None and not usuarios_nuevos_modo.empty:
            usuarios_nuevos_modo.to_excel(writer, sheet_name="Nuevos_MODO", index=False)

        if usuarios_reactivados_modo is not None and not usuarios_reactivados_modo.empty:
            usuarios_reactivados_modo.to_excel(writer, sheet_name="Reactivados_MODO", index=False)
            
        for juego, df in top10_por_juego_con_datos.items():
            if not df.empty:
                df.to_excel(writer, sheet_name=f"Top10_{juego}", index=False)

    print("✅ Hojas de datos guardadas en el Excel.")
    
    # Reabrir el archivo para agregar los gráficos
    wb = load_workbook(SALIDA_ANALITICO)

    if "Resumen" in wb.sheetnames:
        wb.remove(wb["Resumen"])
    ws = wb.create_sheet("Resumen", 0)

    ws["A1"]  = "KPIs principales"
    ws["A3"]  = "Promedio depósito $";            ws["B3"]  = promedio_deposito
    ws["A4"]  = "Usuarios únicos (mov.)";         ws["B4"]  = cant_unicos_total
    ws["A5"]  = "Usuarios únicos apostadores";    ws["B5"]  = cant_unicos_apuestan
    ws["A6"]  = "Usuarios únicos que recargaron"; ws["B6"]  = cant_recargas_unicas
    ws["A8"]  = "Recargas MODO";                  ws["B8"]  = recargas_modo
    ws["A9"]  = "Recargas Retail";                ws["B9"]  = recargas_retail
    ws["A11"] = "Monto MODO $";                   ws["B11"] = monto_modo
    ws["A12"] = "Monto Retail $";                 ws["B12"] = monto_retail
    for cell in ["B3","B11","B12"]:
        ws[cell].number_format = '#,##0.00'
    for cell in ["B4","B5","B6","B8","B9"]:
        ws[cell].number_format = '#,##0'

    # $ por día por canal
    sheet_monto = wb["Recargas_Dia_Monto"]
    max_row = sheet_monto.max_row
    max_col = sheet_monto.max_column
    line1 = LineChart()
    line1.title = "$ por día por canal"
    line1.y_axis.title = "$"
    line1.x_axis.title = "Fecha"
    data_ref = Reference(sheet_monto, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
    cats_ref = Reference(sheet_monto, min_col=1, min_row=2, max_row=max_row)
    line1.add_data(data_ref, titles_from_data=True)
    line1.set_categories(cats_ref)
    line1.height = 11
    line1.width = 24
    ws.add_chart(line1, "D2")

    # Cantidad de recargas por día por canal
    sheet_cnt = wb["Recargas_Dia_Cant"]
    max_row2 = sheet_cnt.max_row
    max_col2 = sheet_cnt.max_column
    bar1 = BarChart()
    bar1.type = "col"
    bar1.title = "Recargas por día por canal"
    bar1.y_axis.title = "Recargas"
    bar1.x_axis.title = "Fecha"
    data_ref2 = Reference(sheet_cnt, min_col=2, min_row=1, max_col=max_col2, max_row=max_row2)
    cats_ref2 = Reference(sheet_cnt, min_col=1, min_row=2, max_row=max_row2)
    bar1.add_data(data_ref2, titles_from_data=True)
    bar1.set_categories(cats_ref2)
    bar1.height = 11
    bar1.width = 24
    ws.add_chart(bar1, "D18")

    # Apuestas por juego (total)
    sheet_games = wb["Top_Games_Total"]
    max_row3 = sheet_games.max_row
    bar2 = BarChart()
    bar2.title = "Apuestas por juego (total)"
    bar2.y_axis.title = "Bets"
    cats3 = Reference(sheet_games, min_col=1, min_row=2, max_row=max_row3)
    data3 = Reference(sheet_games, min_col=2, min_row=1, max_row=max_row3)
    bar2.add_data(data3, titles_from_data=True)
    bar2.set_categories(cats3)
    bar2.height = 11
    bar2.width = 24
    ws.add_chart(bar2, "D34")

    # Apuestas por día de semana
    sheet_dias = wb["Dia_Totales"]
    max_row4 = sheet_dias.max_row
    bar3 = BarChart()
    bar3.title = "Apuestas por día de semana"
    bar3.y_axis.title = "Bets"
    cats4 = Reference(sheet_dias, min_col=1, min_row=2, max_row=max_row4)
    data4 = Reference(sheet_dias, min_col=2, min_row=1, max_row=max_row4)
    bar3.add_data(data4, titles_from_data=True)
    bar3.set_categories(cats4)
    bar3.height = 11
    bar3.width = 24
    ws.add_chart(bar3, "D50")

    # Before vs After 07/07
    sheet_cmp = wb["Comparativa_MODO"]
    bar4 = BarChart()
    bar4.title = f"Before vs After {FECHA_MODO_FULL.strftime('%d/%m/%Y')}"
    bar4.y_axis.title = "$"
    cats5 = Reference(sheet_cmp, min_col=1, min_row=2, max_row=3)
    data5 = Reference(sheet_cmp, min_col=2, min_row=1, max_col=3, max_row=3)
    bar4.add_data(data5, titles_from_data=True)
    bar4.set_categories(cats5)
    bar4.height = 11
    bar4.width = 24
    ws.add_chart(bar4, "D66")

    # Uso por juego por día con filtro de fechas (SUMIFS)
    sheet_det = wb["Juego_Dia_Detalle"]
    min_date = sheet_det["A2"].value
    max_date = sheet_det[f"A{sheet_det.max_row}"].value
    ws["A86"] = "Fecha inicio"; ws["B86"] = min_date
    ws["A87"] = "Fecha fin";    ws["B87"] = max_date
    for c in ("B86","B87"): ws[c].number_format = "yyyy-mm-dd"

    row = 90
    d = pd.to_datetime(min_date).date()
    end = pd.to_datetime(max_date).date()
    while d <= end:
        ws[f"A{row}"] = d
        ws[f"A{row}"].number_format = "yyyy-mm-dd"
        row += 1
        d += timedelta(days=1)
    last_row_dates = row - 1

    max_row_games = sheet_games.max_row
    topN = min(6, max_row_games - 1)
    games = [sheet_games[f"A{r}"].value for r in range(2, 2 + topN)]
    ws["B89"] = "Bets (por juego y día)"
    for idx, g in enumerate(games, start=2):
        ws.cell(row=89, column=idx).value = g

    det_rows = sheet_det.max_row
    for r in range(90, last_row_dates + 1):
        date_cell = f"A{r}"
        for i, g in enumerate(games, start=2):
            ws.cell(row=r, column=i).value = (
                f'=IF(AND({date_cell}>=B86,{date_cell}<=B87),'
                f'SUMIFS(Juego_Dia_Detalle!$C$2:$C${det_rows},'
                f'Juego_Dia_Detalle!$B$2:$B${det_rows},"{g}",'
                f'Juego_Dia_Detalle!$A$2:$A${det_rows},{date_cell}),'
                f'NA())'
            )

    chart = LineChart()
    chart.title = "Uso por juego por día (filtrado)"
    chart.y_axis.title = "Bets"
    chart.x_axis.title = "Fecha"
    min_col = 2
    max_col = 1 + len(games)
    data_ref = Reference(ws, min_col=min_col, min_row=89, max_col=max_col, max_row=last_row_dates)
    cats_ref = Reference(ws, min_col=1, min_row=90, max_row=last_row_dates)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 15
    chart.width  = 28
    ws.add_chart(chart, "D82")

    # Guardar el Excel
    wb.save(SALIDA_ANALITICO)
    print(f"✅ Archivo final guardado con éxito: {SALIDA_ANALITICO}")
except Exception as e:
    print(f"❌ Error al guardar el archivo analítico: {e}")
    # En caso de error, podría ser útil intentar guardar el archivo sin gráficos
    # para no perder la información procesada
    print("Intentando guardar solo los datos...")
    try:
        with pd.ExcelWriter(SALIDA_ANALITICO, engine="openpyxl", mode="w") as writer:
            resumen_kpis.to_excel(writer, sheet_name="Resumen_Datos", index=False)
            cliente_mes.to_excel(writer, sheet_name="Cliente_Mes", index=False)
            top_games_total.to_excel(writer, sheet_name="Top_Games_Total", index=False)
            top_games_mes.to_excel(writer, sheet_name="Top_Games_Mes", index=False)
            recargas_diario_canal.to_excel(writer, sheet_name="Recargas_Diario", index=False)
            recargas_dia_monto.to_excel(writer, sheet_name="Recargas_Dia_Monto", index=False)
            recargas_dia_cant.to_excel(writer, sheet_name="Recargas_Dia_Cant", index=False)
            modo_diario.to_excel(writer, sheet_name="MODO_Diario", index=False)
            retiros_diario.to_excel(writer, sheet_name="Retiros_Diario", index=False)
            premios_resumen.to_excel(writer, sheet_name="Ganadores", index=False)
            juego_dia_detalle.to_excel(writer, sheet_name="Juego_Dia_Detalle", index=False)
            dia_totales.to_excel(writer, sheet_name="Dia_Totales", index=False)
            retencion_modo.to_excel(writer, sheet_name="Retencion_MODO", index=False)
            usuarios_mes.to_excel(writer, sheet_name="Usuarios_Mes", index=False)
            usuarios_hitos.to_excel(writer, sheet_name="Usuarios_Hitos", index=False)
            comparativa_modo.to_excel(writer, sheet_name="Comparativa_MODO", index=False)
            if top10_contactos is not None and not top10_contactos.empty:
                top10_contactos.to_excel(writer, sheet_name="Top10_Contactos", index=False)
            if usuarios_nuevos_modo is not None and not usuarios_nuevos_modo.empty:
                usuarios_nuevos_modo.to_excel(writer, sheet_name="Nuevos_MODO", index=False)
            if usuarios_reactivados_modo is not None and not usuarios_reactivados_modo.empty:
                usuarios_reactivados_modo.to_excel(writer, sheet_name="Reactivados_MODO", index=False)
            for juego, df in top10_por_juego_con_datos.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=f"Top10_{juego}", index=False)
        print("✅ Archivo guardado sin gráficos.")
    except Exception as e_retry:
        print(f"❌ Error crítico al intentar guardar de nuevo: {e_retry}. No se pudo generar el reporte.")

# --- Exportar CSVs DIRECTO desde DataFrames en memoria ---
try:
    modo_diario.to_csv(csv_dir / "modo_diario.csv", index=False)
    recargas_dia_monto.to_csv(csv_dir / "recargas_monto.csv", index=False)
    recargas_dia_cant.to_csv(csv_dir / "recargas_cant.csv", index=False)
    comparativa_modo.to_csv(csv_dir / "comparativa_modo.csv", index=False)
    resumen_kpis.to_csv(csv_dir / "kpis.csv", index=False)

    # (Opcional) si necesitás un CSV por juego/mes
    top_games_mes.to_csv(csv_dir / "jugadores_unicos_por_juego.csv", index=False)

    # deposito_promedio.csv SIN re-abrir Excel:
    prom = recargas_dia_monto.copy()
    for col in ["MODO", "Retail"]:
        if col in recargas_dia_monto.columns and col in recargas_dia_cant.columns:
            prom[col] = recargas_dia_monto[col] / recargas_dia_cant[col]
    prom.to_csv(csv_dir / "deposito_promedio.csv", index=False)

    print("✅ CSVs escritos directamente desde memoria.")
    print(f"📁 Carpeta CSV: {csv_dir.resolve()}")
except Exception as e:
    print(f"⚠️ Error al exportar CSVs desde memoria: {e}")


# ========= Exportar reportes adicionales (usuarios) =========
try:
    excel_file = pd.ExcelFile(SALIDA_ANALITICO)

    if "Nuevos_MODO" in excel_file.sheet_names:
        df = excel_file.parse("Nuevos_MODO")
        df.to_csv(csv_dir / "nuevos_modo.csv", index=False)

    if "Reactivados_MODO" in excel_file.sheet_names:
        df = excel_file.parse("Reactivados_MODO")
        df.to_csv(csv_dir / "reactivados_modo.csv", index=False)

    for juego, df in top10_por_juego_con_datos.items():
        archivo = f"top10_{juego.lower().replace(' ', '_')}.csv"
        if not df.empty:
            df.to_csv(csv_dir / archivo, index=False)

    print("✅ CSV adicionales exportados (usuarios únicos, nuevos, reactivados, top10 por juego)")
except Exception as e:
    print(f"⚠️ Error al exportar reportes adicionales: {e}")

print("🏁  Proceso incremental terminado.")