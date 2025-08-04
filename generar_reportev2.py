#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
import warnings
import unicodedata
from datetime import datetime, timedelta
import pandas as pd
import numpy as np

# ========= CONFIG =========
DATA_DIR          = Path("data")                       # carpeta con .xls
SALIDA_ANALITICO  = "reporte_movimientos.xlsx"         # archivo con datos + gráficos openpyxl
USUARIOS_FILE     = Path("usuarios.xlsx")              # opcional
TEMPLATE_PATH     = Path("Plantilla-Reporte.xlsm")     # plantilla de pivots/slicers
SALIDA_XLWINGS    = f"ReporteMovimientos-{datetime.now():%Y%m%d}.xlsx"
ENABLE_XLWINGS    = True   # ← poné False si no querés empujar a la plantilla

# Hitos funcionales
FECHA_LANZ_JUEGOS = pd.Timestamp("2025-04-14")         # Quini6 + Loto Plus
FECHA_MODO_FULL   = pd.Timestamp("2025-07-07")         # MODO disponible para todos

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style",
    module="openpyxl.styles.stylesheet",
)

# ========= UTILIDADES =========
def normalizar(txt: str) -> str:
    if pd.isna(txt):
        return ""
    txt = unicodedata.normalize("NFKD", str(txt))
    return "".join(c for c in txt if not unicodedata.combining(c)).lower()

def leer_movimientos(archivo: Path) -> pd.DataFrame | None:
    crudo = pd.read_excel(archivo, header=None)
    header_mask = crudo.apply(
        lambda fila: fila.apply(normalizar).str.contains("tipo mov", na=False).any(),
        axis=1,
    )
    if not header_mask.any():
        print(f"⚠️  Encabezado no encontrado en {archivo.name} — se omite")
        return None
    header_idx = header_mask.idxmax()

    df = pd.read_excel(
        archivo,
        header=header_idx,
        usecols=[
            "Nro. Transacción",
            "Fecha",
            "Tipo Mov.",
            "Documento",
            "Movimiento",
            "Importe",
        ],
    )
    df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors="coerce")
    df["Importe"] = (
        df["Importe"].astype(str)
                     .str.replace(".", "", regex=False)   # miles
                     .str.replace(",", ".", regex=False)  # decimales
                     .astype(float)
    )
    return df

def clasificar_canal(movimiento: str) -> str:
    # Solo dos categorías
    return "MODO" if "modo" in normalizar(movimiento) else "Retail"

# ========= CARGA =========
archivos = sorted(f for f in DATA_DIR.glob("*.xls") if not f.name.startswith("~$"))
dfs = []
for f in archivos:
    try:
        tmp = leer_movimientos(f)
    except FileNotFoundError:
        print(f"⚠️  {f.name} ya no está disponible, se omite.")
        continue
    if tmp is not None:
        dfs.append(tmp)

if not dfs:
    raise RuntimeError("No se encontró ningún archivo válido en 'data/'.")

data = pd.concat(dfs, ignore_index=True)

# ========= APUESTAS =========
apuestas = data[data["Tipo Mov."].str.contains("apuesta|jugada", case=False, na=False)].copy()
apuestas["AñoMes"]    = apuestas["Fecha"].dt.to_period("M")
apuestas["Fecha_Dia"] = apuestas["Fecha"].dt.date
apuestas["Juego"] = (
    apuestas["Movimiento"].str.replace(r"(?i)jugada\s*-\s*", "", regex=True).str.strip()
)
apuestas["Juego_norm"] = apuestas["Juego"].apply(normalizar)

dias_map = {0:"Lunes",1:"Martes",2:"Miércoles",3:"Jueves",4:"Viernes",5:"Sábado",6:"Domingo"}
apuestas["Dia_Sem"] = apuestas["Fecha"].dt.weekday.map(dias_map)

# ========= CLIENTE_MES =========
cliente_mes = (
    apuestas.groupby(["Documento", "AñoMes", "Juego"])
            .agg(Bets=("Importe", "count"), Gastado=("Importe", "sum"))
            .reset_index()
)

# ========= TOP GAMES =========
top_games_total = (
    apuestas.groupby("Juego")
            .agg(Bets_Totales=("Importe", "count"),
                 Jugadores_Unicos=("Documento", "nunique"),
                 Gastado=("Importe", "sum"))
            .sort_values("Bets_Totales", ascending=False)
            .reset_index()
)

top_games_mes = (
    apuestas.groupby(["AñoMes", "Juego"])
            .agg(Bets_Mes=("Importe", "count"),
                 Jugadores_Mes=("Documento", "nunique"),
                 Gastado_Mes=("Importe", "sum"))
            .reset_index()
            .sort_values(["AñoMes", "Bets_Mes"], ascending=[False, False])
)

# ========= PESTAÑAS POR JUEGO =========
GAME_PATTERNS = {
    "Tombo_Express": r"tombo express",
    "Tombola":       r"t(?:o|ó)mbola",
    "Quini6":        r"quini\s*6",
    "Loto_Plus":     r"loto(?:\s*plus)?",
}
game_summaries = {}
for sheet, pattern in GAME_PATTERNS.items():
    mask = apuestas["Juego_norm"].str.contains(pattern, regex=True, na=False)
    tmp = apuestas.loc[mask]
    summary = (
        tmp.groupby("Documento")
           .agg(Bets=("Importe", "count"), Gastado=("Importe", "sum"))
           .sort_values("Gastado", ascending=False)
           .reset_index()
    )
    game_summaries[sheet] = summary

# ========= RECARGAS / RETIROS / PREMIOS =========
mask_carga = data["Tipo Mov."].str.contains(
    r"carga|dep(?:o|ó)sito", case=False, regex=True, na=False
)
cargas = data.loc[mask_carga].copy()
cargas["Canal"]     = cargas["Movimiento"].apply(clasificar_canal)
cargas["Metodo"]    = np.where(cargas["Canal"]=="MODO", "MODO", "Retail")
cargas["Fecha_Dia"] = cargas["Fecha"].dt.date
cargas["Hora"]      = cargas["Fecha"].dt.hour

recargas_diario_canal = (
    cargas.groupby(["Fecha_Dia", "Canal"])
          .agg(Recargas=("Importe", "count"),
               Monto=("Importe", "sum"),
               Usuarios_Unicos=("Documento", "nunique"))
          .reset_index()
)

recargas_dia_monto = (
    recargas_diario_canal.pivot(index="Fecha_Dia", columns="Canal", values="Monto")
                         .fillna(0.0)
                         .reset_index()
                         .sort_values("Fecha_Dia")
)
recargas_dia_cant = (
    recargas_diario_canal.pivot(index="Fecha_Dia", columns="Canal", values="Recargas")
                         .fillna(0)
                         .reset_index()
                         .sort_values("Fecha_Dia")
)

modo_diario = (
    cargas.loc[cargas["Canal"] == "MODO"]
          .groupby("Fecha_Dia")
          .agg(
              Recargas_MODO   = ("Importe", "count"),
              Monto_MODO      = ("Importe", "sum"),
              Usuarios_Unicos = ("Documento", "nunique"),
          )
          .reset_index()
          .sort_values("Fecha_Dia", ascending=False)
)

mask_retiro = data["Tipo Mov."].str.contains("retiro|transferencia salida", case=False, na=False)
retiros = data.loc[mask_retiro].copy()
retiros["Fecha_Dia"] = retiros["Fecha"].dt.date
retiros_diario = (
    retiros.groupby("Fecha_Dia")
           .agg(Retiros=("Importe", "count"),
                Monto_Retirado=("Importe", "sum"),
                Clientes_Unicos=("Documento", "nunique"))
           .reset_index()
           .sort_values("Fecha_Dia", ascending=False)
)

mask_premio = data["Tipo Mov."].str.contains("premio", case=False, na=False)
premios = data.loc[mask_premio].copy()
premios_resumen = (
    premios.groupby("Documento")
           .agg(Premios_Cobrados=("Importe", "count"), Monto_Premios=("Importe", "sum"))
           .sort_values("Monto_Premios", ascending=False)
           .reset_index()
)

# ========= Juego por día (detalle) =========
juego_dia_detalle = (
    apuestas.groupby(["Fecha_Dia", "Juego"])
            .agg(Bets=("Importe", "count"),
                 Usuarios_Unicos_Dia=("Documento", "nunique"),
                 Gastado_Dia=("Importe", "sum"))
            .reset_index()
            .sort_values(["Fecha_Dia", "Juego"])
)

dia_totales = (
    apuestas.groupby("Dia_Sem")
            .agg(Bets=("Importe", "count"),
                 Usuarios_Unicos=("Documento", "nunique"))
            .reset_index()
)

# ========= Retención MODO =========
primer_mov_total = (
    data.sort_values("Fecha")
        .groupby("Documento", as_index=False)
        .first()
        .rename(columns={"Fecha": "Fecha_PrimerMov"})
)

modo_all = cargas.loc[cargas["Canal"] == "MODO"].copy()
primera_modo = (
    modo_all.sort_values("Fecha")
            .groupby("Documento", as_index=False)
            .first()
            .rename(columns={"Fecha": "Fecha_Corte"})
)
retencion_base = primera_modo.merge(primer_mov_total, on="Documento", how="left")
retencion_base["Es_Nuevo"] = retencion_base["Fecha_PrimerMov"] == retencion_base["Fecha_Corte"]

apuestas_only = apuestas[["Documento", "Fecha"]].copy()
joined = retencion_base[["Documento", "Fecha_Corte"]].merge(apuestas_only, on="Documento", how="left")
joined["Posterior"]     = joined["Fecha"] >  joined["Fecha_Corte"]
joined["Dia_Siguiente"] = (joined["Fecha"] > joined["Fecha_Corte"]) & \
                          (joined["Fecha"] <= joined["Fecha_Corte"] + pd.Timedelta(days=1))
joined["Mes_Siguiente"] = (joined["Fecha"] > joined["Fecha_Corte"]) & \
                          (joined["Fecha"] <= joined["Fecha_Corte"] + pd.Timedelta(days=30))

flags = (
    joined.groupby("Documento")
          .agg(Jugo_Posterior=("Posterior", "any"),
               Jugo_Dia_Sig=("Dia_Siguiente", "any"),
               Jugo_Mes_Sig=("Mes_Siguiente", "any"))
          .reset_index()
)
retencion_modo = retencion_base.merge(flags, on="Documento", how="left").fillna(False)
print(f"▶︎ Cantidad de usuarios NUEVOS que cargaron con MODO: {int(retencion_modo['Es_Nuevo'].sum())}")

# ========= Crecimiento / Hitos =========
primer_mov_total["AñoMes_PrimerMov"] = primer_mov_total["Fecha_PrimerMov"].dt.to_period("M")
usuarios_mes = (
    primer_mov_total.groupby("AñoMes_PrimerMov")
                    .agg(Nuevos=("Documento", "nunique"))
                    .reset_index()
                    .sort_values("AñoMes_PrimerMov")
)
usuarios_mes["Acumulado"] = usuarios_mes["Nuevos"].cumsum()
activos_mes = (
    apuestas.groupby(apuestas["Fecha"].dt.to_period("M"))["Documento"]
            .nunique()
            .reset_index(name="Jugadores_Activos_Mes")
            .rename(columns={"Fecha": "AñoMes"})
)
activos_mes["AñoMes"] = activos_mes["AñoMes"].astype(str)
usuarios_mes["AñoMes"] = usuarios_mes["AñoMes_PrimerMov"].astype(str)
usuarios_mes = usuarios_mes.merge(activos_mes, on="AñoMes", how="left")
usuarios_mes = usuarios_mes[["AñoMes", "Nuevos", "Acumulado", "Jugadores_Activos_Mes"]]

nuevos_desde_juegos = primer_mov_total.loc[
    primer_mov_total["Fecha_PrimerMov"] >= FECHA_LANZ_JUEGOS, "Documento"
].nunique()
jugadores_desde_juegos = apuestas.loc[
    apuestas["Fecha"] >= FECHA_LANZ_JUEGOS, "Documento"
].nunique()
mask_newgames = (
    (apuestas["Fecha"] >= FECHA_LANZ_JUEGOS) &
    apuestas["Juego_norm"].str.contains(r"(?:quini\s*6|loto(?:\s*plus)?)", regex=True, na=False)
)
jugadores_quini_loto_desde = apuestas.loc[mask_newgames, "Documento"].nunique()

modo_post_docs = set(
    cargas.loc[(cargas["Fecha"] >= FECHA_MODO_FULL) & (cargas["Canal"] == "MODO"), "Documento"].unique()
)
usuarios_modo_desde = len(modo_post_docs)
jugadores_post_modo_docs = set(apuestas.loc[apuestas["Fecha"] >= FECHA_MODO_FULL, "Documento"].unique())
jugadores_post_modo = len(jugadores_post_modo_docs)
jugadores_modo_y_jugaron = len(modo_post_docs & jugadores_post_modo_docs)

usuarios_hitos = pd.DataFrame({
    "Concepto": [
        "Nuevos >= 2025-04-14 (cualquier mov.)",
        "Jugadores >= 2025-04-14 (cualquier apuesta)",
        "Apostaron Quini/Loto >= 2025-04-14",
        "Recargaron MODO >= 2025-07-07",
        "Jugadores >= 2025-07-07 (cualquier apuesta)",
        "Recargaron MODO y jugaron >= 2025-07-07",
    ],
    "Valor": [
        nuevos_desde_juegos,
        jugadores_desde_juegos,
        jugadores_quini_loto_desde,
        usuarios_modo_desde,
        jugadores_post_modo,
        jugadores_modo_y_jugaron,
    ],
})

# ========= Comparativa Before/After 07/07 =========
dep_before = cargas.loc[cargas["Fecha"] < FECHA_MODO_FULL, "Importe"].sum()
dep_after  = cargas.loc[cargas["Fecha"] >= FECHA_MODO_FULL, "Importe"].sum()
rec_before = apuestas.loc[apuestas["Fecha"] < FECHA_MODO_FULL, "Importe"].sum()
rec_after  = apuestas.loc[apuestas["Fecha"] >= FECHA_MODO_FULL, "Importe"].sum()

comparativa_modo = pd.DataFrame({
    "Periodo": ["Before 07/07/2025", "After 07/07/2025"],
    "Depositos_$": [dep_before, dep_after],
    "Recaudacion_$": [rec_before, rec_after],
})

# ========= KPIs Resumen =========
promedio_deposito    = cargas["Importe"].mean() if len(cargas) else 0.0
cant_unicos_total    = data["Documento"].nunique()
cant_unicos_apuestan = apuestas["Documento"].nunique()
cant_recargas_unicas = cargas["Documento"].nunique()

agg_canal = (
    cargas.groupby("Canal")
          .agg(Recargas=("Importe", "count"),
               Usuarios_Unicos=("Documento", "nunique"),
               Monto=("Importe", "sum"))
          .reset_index()
)
recargas_modo   = int(agg_canal.loc[agg_canal["Canal"]=="MODO","Recargas"].sum())
recargas_retail = int(agg_canal.loc[agg_canal["Canal"]=="Retail","Recargas"].sum())
monto_modo      = float(agg_canal.loc[agg_canal["Canal"]=="MODO","Monto"].sum())
monto_retail    = float(agg_canal.loc[agg_canal["Canal"]=="Retail","Monto"].sum())

resumen_kpis = pd.DataFrame({
    "KPI": [
        "Promedio depósito $",
        "Usuarios únicos (cualquier mov.)",
        "Usuarios únicos apostadores",
        "Usuarios únicos que recargaron",
        "Recargas - MODO",
        "Recargas - Retail",
        "Monto MODO $",
        "Monto Retail $",
    ],
    "Valor": [
        promedio_deposito,
        cant_unicos_total,
        cant_unicos_apuestan,
        cant_recargas_unicas,
        recargas_modo,
        recargas_retail,
        monto_modo,
        monto_retail,
    ]
})

# ========= Top10 contactos (opcional) =========
top10_contactos = None
if USUARIOS_FILE.exists():
    try:
        usuarios = pd.read_excel(USUARIOS_FILE)
        usuarios.columns = usuarios.columns.str.strip()
        if "DNI" not in usuarios.columns:
            cand = [c for c in usuarios.columns if "dni" in c.lower() or "doc" in c.lower()]
            if cand:
                usuarios.rename(columns={cand[0]: "DNI"}, inplace=True)
            else:
                raise KeyError("No se encontró columna DNI en usuarios.xlsx")
        usuarios["DNI"] = pd.to_numeric(usuarios["DNI"], errors="coerce")
        usuarios = usuarios.dropna(subset=["DNI"]).drop_duplicates(subset="DNI", keep="last")

        jugadas_por_doc = (
            apuestas.groupby("Documento")
                    .agg(Bets_Total=("Importe", "count"),
                         Gastado_Total=("Importe", "sum"))
                    .reset_index()
        )

        top10_contactos = (
            jugadas_por_doc.merge(usuarios, how="inner", left_on="Documento", right_on="DNI")
                           .sort_values("Bets_Total", ascending=False)
                           .head(10)
        )
    except Exception as e:
        print(f"⚠️  No se pudo generar Top10_Contactos: {e}")

# ========= Exportar: archivo analítico con gráficos (openpyxl) =========
with pd.ExcelWriter(SALIDA_ANALITICO, engine="openpyxl", mode="w") as writer:
    resumen_kpis.to_excel(writer,              sheet_name="Resumen_Datos",     index=False)
    cliente_mes.to_excel(writer,               sheet_name="Cliente_Mes",       index=False)
    top_games_total.to_excel(writer,           sheet_name="Top_Games_Total",   index=False)
    top_games_mes.to_excel(writer,             sheet_name="Top_Games_Mes",     index=False)

    for sheet, df in game_summaries.items():
        df.to_excel(writer, sheet_name=sheet, index=False)

    recargas_diario_canal.to_excel(writer,     sheet_name="Recargas_Diario",   index=False)
    recargas_dia_monto.to_excel(writer,        sheet_name="Recargas_Dia_Monto",index=False)
    recargas_dia_cant.to_excel(writer,         sheet_name="Recargas_Dia_Cant", index=False)
    modo_diario.to_excel(writer,               sheet_name="MODO_Diario",       index=False)
    retiros_diario.to_excel(writer,            sheet_name="Retiros_Diario",    index=False)
    premios_resumen.to_excel(writer,           sheet_name="Ganadores",         index=False)
    juego_dia_detalle.to_excel(writer,         sheet_name="Juego_Dia_Detalle", index=False)
    dia_totales.to_excel(writer,               sheet_name="Dia_Totales",       index=False)
    retencion_modo.to_excel(writer,            sheet_name="Retencion_MODO",    index=False)
    usuarios_mes.to_excel(writer,              sheet_name="Usuarios_Mes",      index=False)
    usuarios_hitos.to_excel(writer,            sheet_name="Usuarios_Hitos",    index=False)
    comparativa_modo.to_excel(writer,          sheet_name="Comparativa_MODO",  index=False)

    if top10_contactos is not None:
        top10_contactos.to_excel(writer,      sheet_name="Top10_Contactos",    index=False)

# ----- Hoja RESUMEN + gráficos
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference

wb = load_workbook(SALIDA_ANALITICO)
if "Resumen" in wb.sheetnames:
    wb.remove(wb["Resumen"])
ws = wb.create_sheet("Resumen", 0)

ws["A1"]  = "KPIs principales"
ws["A3"]  = "Promedio depósito $";            ws["B3"]  = promedio_deposito
ws["A4"]  = "Usuarios únicos (mov.)";         ws["B4"]  = cant_unicos_total
ws["A5"]  = "Usuarios únicos apostadores";    ws["B5"]  = cant_unicos_apuestan
ws["A6"]  = "Usuarios únicos que recargaron"; ws["B6"]  = cant_recargas_unicas
ws["A8"]  = "Recargas MODO";                  ws["B8"]  = recargas_modo
ws["A9"]  = "Recargas Retail";                ws["B9"]  = recargas_retail
ws["A11"] = "Monto MODO $";                   ws["B11"] = monto_modo
ws["A12"] = "Monto Retail $";                 ws["B12"] = monto_retail
for cell in ["B3","B11","B12"]:
    ws[cell].number_format = '#,##0.00'
for cell in ["B4","B5","B6","B8","B9"]:
    ws[cell].number_format = '#,##0'

# $ por día por canal
sheet_monto = wb["Recargas_Dia_Monto"]
max_row = sheet_monto.max_row
max_col = sheet_monto.max_column
line1 = LineChart()
line1.title = "$ por día por canal"
line1.y_axis.title = "$"
line1.x_axis.title = "Fecha"
data_ref = Reference(sheet_monto, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
cats_ref = Reference(sheet_monto, min_col=1, min_row=2, max_row=max_row)
line1.add_data(data_ref, titles_from_data=True)
line1.set_categories(cats_ref)
line1.height = 11
line1.width = 24
ws.add_chart(line1, "D2")

# Cantidad de recargas por día por canal
sheet_cnt = wb["Recargas_Dia_Cant"]
max_row2 = sheet_cnt.max_row
max_col2 = sheet_cnt.max_column
bar1 = BarChart()
bar1.type = "col"
bar1.title = "Recargas por día por canal"
bar1.y_axis.title = "Recargas"
bar1.x_axis.title = "Fecha"
data_ref2 = Reference(sheet_cnt, min_col=2, min_row=1, max_col=max_col2, max_row=max_row2)
cats_ref2 = Reference(sheet_cnt, min_col=1, min_row=2, max_row=max_row2)
bar1.add_data(data_ref2, titles_from_data=True)
bar1.set_categories(cats_ref2)
bar1.height = 11
bar1.width = 24
ws.add_chart(bar1, "D18")

# Apuestas por juego (total)
sheet_games = wb["Top_Games_Total"]
max_row3 = sheet_games.max_row
bar2 = BarChart()
bar2.title = "Apuestas por juego (total)"
bar2.y_axis.title = "Bets"
cats3 = Reference(sheet_games, min_col=1, min_row=2, max_row=max_row3)
data3 = Reference(sheet_games, min_col=2, min_row=1, max_row=max_row3)
bar2.add_data(data3, titles_from_data=True)
bar2.set_categories(cats3)
bar2.height = 11
bar2.width = 24
ws.add_chart(bar2, "D34")

# Apuestas por día de semana
sheet_dias = wb["Dia_Totales"]
max_row4 = sheet_dias.max_row
bar3 = BarChart()
bar3.title = "Apuestas por día de semana"
bar3.y_axis.title = "Bets"
cats4 = Reference(sheet_dias, min_col=1, min_row=2, max_row=max_row4)
data4 = Reference(sheet_dias, min_col=2, min_row=1, max_row=max_row4)
bar3.add_data(data4, titles_from_data=True)
bar3.set_categories(cats4)
bar3.height = 11
bar3.width = 24
ws.add_chart(bar3, "D50")

# Before vs After 07/07
sheet_cmp = wb["Comparativa_MODO"]
bar4 = BarChart()
bar4.title = "Before vs After 07/07/2025"
bar4.y_axis.title = "$"
cats5 = Reference(sheet_cmp, min_col=1, min_row=2, max_row=3)
data5 = Reference(sheet_cmp, min_col=2, min_row=1, max_col=3, max_row=3)
bar4.add_data(data5, titles_from_data=True)
bar4.set_categories(cats5)
bar4.height = 11
bar4.width = 24
ws.add_chart(bar4, "D66")

# Uso por juego por día con filtro de fechas (SUMIFS)
sheet_det = wb["Juego_Dia_Detalle"]
min_date = sheet_det["A2"].value
max_date = sheet_det[f"A{sheet_det.max_row}"].value
ws["A86"] = "Fecha inicio"; ws["B86"] = min_date
ws["A87"] = "Fecha fin";    ws["B87"] = max_date
for c in ("B86","B87"): ws[c].number_format = "yyyy-mm-dd"

row = 90
d = pd.to_datetime(min_date).date()
end = pd.to_datetime(max_date).date()
while d <= end:
    ws[f"A{row}"] = d
    ws[f"A{row}"].number_format = "yyyy-mm-dd"
    row += 1
    d += timedelta(days=1)
last_row_dates = row - 1

max_row_games = sheet_games.max_row
topN = min(6, max_row_games - 1)
games = [sheet_games[f"A{r}"].value for r in range(2, 2 + topN)]
ws["B89"] = "Bets (por juego y día)"
for idx, g in enumerate(games, start=2):
    ws.cell(row=89, column=idx).value = g

det_rows = sheet_det.max_row
for r in range(90, last_row_dates + 1):
    date_cell = f"A{r}"
    for i, g in enumerate(games, start=2):
        ws.cell(row=r, column=i).value = (
            f'=IF(AND({date_cell}>=B86,{date_cell}<=B87),'
            f'SUMIFS(Juego_Dia_Detalle!$C$2:$C${det_rows},'
            f'Juego_Dia_Detalle!$B$2:$B${det_rows},"{g}",'
            f'Juego_Dia_Detalle!$A$2:$A${det_rows},{date_cell}),'
            f'NA())'
        )

from openpyxl.chart import LineChart
chart = LineChart()
chart.title = "Uso por juego por día (filtrado)"
chart.y_axis.title = "Bets"
chart.x_axis.title = "Fecha"
min_col = 2
max_col = 1 + len(games)
data_ref = Reference(ws, min_col=min_col, min_row=89, max_col=max_col, max_row=last_row_dates)
cats_ref = Reference(ws, min_col=1, min_row=90, max_row=last_row_dates)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 15
chart.width  = 28
ws.add_chart(chart, "D82")

wb.save(SALIDA_ANALITICO)
print(f"✅ Reporte con gráficos generado: {SALIDA_ANALITICO}")

# ========= xlwings / Plantilla (opcional) =========
if ENABLE_XLWINGS:
    try:
        import xlwings as xw
        from pandas.api.types import is_period_dtype

        def inspeccionar_plantilla(path: Path):
            app = xw.App(visible=False)
            try:
                wb = app.books.open(str(path))
                print("---- Tablas detectadas ----")
                for sht in wb.sheets:
                    los = sht.api.ListObjects
                    if los.Count > 0:
                        for lo in los:
                            h = lo.HeaderRowRange
                            d = lo.DataBodyRange
                            print(f"Hoja={sht.name:20} Tabla={lo.Name:20} "
                                  f"Header={h.Address} Body={(d.Address if d is not None else '(sin datos)')}")
                wb.close()
            finally:
                app.quit()

        # Descomentar para inspeccionar la plantilla una vez:
        # inspeccionar_plantilla(TEMPLATE_PATH)

        def assert_table_exists(ws, table_name: str):
            names = [lo.Name for lo in ws.api.ListObjects]
            if table_name not in names:
                raise RuntimeError(
                    f"En hoja '{ws.name}' no existe la tabla '{table_name}'. "
                    f"Tablas encontradas: {names}"
                )

        def write_df_to_table(ws, table_name: str, df: pd.DataFrame):
            """
            Escribe df en la tabla: posiciona por Row/Column del encabezado,
            convierte Period->str, escribe encabezados+datos y redimensiona la ListObject.
            """
            df2 = df.copy()
            # Period → str
            for c in df2.columns:
                if is_period_dtype(df2[c]):
                    df2[c] = df2[c].astype(str)
            # NaN → ""
            df2 = df2.where(df2.notna(), "")
            nrows, ncols = df2.shape

            tbl = ws.api.ListObjects(table_name)
            tl = tbl.HeaderRowRange.Cells(1, 1)

            start_row = tl.Row
            start_col = tl.Column

            # Encabezados
            ws.range((start_row, start_col)).value = df2.columns.tolist()
            # Datos
            if nrows > 0:
                ws.range((start_row + 1, start_col)).value = df2.values.tolist()

            # Rango final (incluye encabezado)
            last_row = start_row + nrows
            last_col = start_col + ncols - 1
            final_rng = ws.range((start_row, start_col), (last_row, last_col))

            # Redimensionar tabla
            tbl.Resize(final_rng.api)

        def actualizar_plantilla(template_path: Path, salida_path: Path,
                                 df_apuestas: pd.DataFrame,
                                 df_cargas: pd.DataFrame,
                                 df_retiros: pd.DataFrame,
                                 df_premios: pd.DataFrame,
                                 df_juego_dia: pd.DataFrame,
                                 df_modo_diario: pd.DataFrame,
                                 df_comparativa: pd.DataFrame,
                                 visible=False):
            app = xw.App(visible=visible)
            app.display_alerts = False
            app.screen_updating = False
            try:
                wb = app.books.open(str(template_path))

                ws_apu = wb.sheets["datos_apuestas"]
                ws_car = wb.sheets["datos_cargas"]
                ws_ret = wb.sheets["datos_retiros"]
                ws_pre = wb.sheets["datos_premios"]
                ws_jdd = wb.sheets["datos_juego_dia"]
                ws_mod = wb.sheets["datos_modo_diario"]
                ws_cmp = wb.sheets["datos_comparativa"]

                # Verificación de tablas
                assert_table_exists(ws_apu, "tblApuestas")
                assert_table_exists(ws_car, "tblCargas")
                assert_table_exists(ws_ret, "tblRetiros")
                assert_table_exists(ws_pre, "tblPremios")
                assert_table_exists(ws_jdd, "tblJuegoDia")
                assert_table_exists(ws_mod, "tblModoDiario")
                assert_table_exists(ws_cmp, "tblComparativaModo")

                # Asegurar AñoMes como str (redundante con write_df_to_table, pero seguro)
                df_apuestas = df_apuestas.copy()
                if is_period_dtype(df_apuestas["AñoMes"]):
                    df_apuestas["AñoMes"] = df_apuestas["AñoMes"].astype(str)

                # Escribir datos
                write_df_to_table(ws_apu, "tblApuestas",
                                  df_apuestas[["Fecha","Fecha_Dia","AñoMes","Documento","Juego","Importe"]])
                write_df_to_table(ws_car, "tblCargas",
                                  df_cargas[["Fecha","Fecha_Dia","Hora","Documento","Metodo","Canal","Importe"]])
                write_df_to_table(ws_ret, "tblRetiros",
                                  df_retiros[["Fecha","Fecha_Dia","Documento","Importe"]])
                write_df_to_table(ws_pre, "tblPremios",
                                  df_premios[["Fecha","Documento","Importe"]])
                write_df_to_table(ws_jdd, "tblJuegoDia",
                                  df_juego_dia[["Fecha_Dia","Juego","Bets","Usuarios_Unicos_Dia","Gastado_Dia"]])
                write_df_to_table(ws_mod, "tblModoDiario",
                                  df_modo_diario[["Fecha_Dia","Recargas_MODO","Monto_MODO","Usuarios_Unicos"]])
                write_df_to_table(ws_cmp, "tblComparativaModo",
                                  df_comparativa[["Periodo","Depositos_$","Recaudacion_$"]])

                # Refrescar pivots
                for pc in wb.api.PivotCaches():
                    try:
                        pc.Refresh()
                    except Exception as e:
                        print("PivotCache refresh error:", e)

                wb.save(str(salida_path))
                wb.close()
            finally:
                app.quit()

        # Verificación mínima de columnas
        req_cols = {"Fecha","Fecha_Dia","AñoMes","Documento","Juego","Importe"}
        faltan = req_cols - set(apuestas.columns)
        if faltan:
            raise RuntimeError(f"Faltan columnas en 'apuestas' para plantilla: {faltan}")

        if not TEMPLATE_PATH.exists():
            print(f"⚠️  Plantilla no encontrada: {TEMPLATE_PATH}. Omite xlwings.")
        else:
            actualizar_plantilla(
                TEMPLATE_PATH, SALIDA_XLWINGS,
                df_apuestas   = apuestas,
                df_cargas     = cargas,
                df_retiros    = retiros,
                df_premios    = premios,
                df_juego_dia  = juego_dia_detalle,
                df_modo_diario= modo_diario,
                df_comparativa= comparativa_modo,
                visible=False,   # poné True para ver Excel en pantalla
            )
            print(f"✅ Archivo con pivots actualizado: {SALIDA_XLWINGS}")

    except ImportError:
        print("ℹ️  xlwings no está instalado. Ejecutá: pip install xlwings")
